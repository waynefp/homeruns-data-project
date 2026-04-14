"""
Create the HR Tracking Excel workbook with all 5 tabs, pre-populated with backfill data.
Upload the resulting .xlsx to Google Drive to use as a Google Sheet.
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data"
OUTPUT_FILE = Path(__file__).parent / "HR_Tracker_2026.xlsx"

# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FILL_ALT = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row_num, fill, num_cols):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=30):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def load_csv(filename):
    """Load a CSV file and return headers + rows."""
    filepath = DATA_DIR / filename
    if not filepath.exists():
        return [], []
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = list(reader)
    return headers, rows


def create_roster_sheet(wb):
    """Sheet 1: Player Roster."""
    ws = wb.active
    ws.title = "Roster"

    headers = [
        "Player Name", "Player ID", "Team", "League", "Bats",
        "2025 HRs", "Status", "2026 HRs (auto)", "2026 Games (auto)",
        "2026 HR Rate",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, HEADER_FILL, len(headers))

    csv_headers, rows = load_csv("roster.csv")
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row):
            cell = ws.cell(row=i, column=j + 1)
            # Try to convert to int where appropriate
            if j in (1, 5):  # player_id, hr_2025
                try:
                    cell.value = int(val)
                except ValueError:
                    cell.value = val
            else:
                cell.value = val
            cell.border = THIN_BORDER

        # Formula columns: 2026 HRs (sum from Daily Log), Games, HR Rate
        r = i
        # SUMPRODUCT to count HRs from Daily Game Log where player matches
        ws.cell(row=r, column=8, value=f'=SUMPRODUCT(("Daily Game Log"!B$2:B$500=A{r})*"Daily Game Log"!J$2:J$500)')
        ws.cell(row=r, column=9, value=f'=COUNTIF("Daily Game Log"!B$2:B$500,A{r})')
        ws.cell(row=r, column=10, value=f'=IF(I{r}>0,H{r}/I{r},"")')
        ws.cell(row=r, column=8).border = THIN_BORDER
        ws.cell(row=r, column=9).border = THIN_BORDER
        ws.cell(row=r, column=10).border = THIN_BORDER
        ws.cell(row=r, column=10).number_format = '0.000'

    auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


def create_daily_log_sheet(wb):
    """Sheet 2: Daily Game Log."""
    ws = wb.create_sheet("Daily Game Log")

    headers = [
        "Date", "Player Name", "Player ID", "Team", "League",
        "Opponent", "Home/Away", "AB", "Hits", "HRs",
        "RBI", "BB", "SO", "Game PK",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, HEADER_FILL_ALT, len(headers))

    csv_headers, rows = load_csv("daily_game_log.csv")
    int_cols = {7, 8, 9, 10, 11, 12, 13}  # AB through game_pk (0-indexed: 7-13)
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row):
            cell = ws.cell(row=i, column=j + 1)
            if j in int_cols:
                try:
                    cell.value = int(val)
                except ValueError:
                    cell.value = val
            else:
                cell.value = val
            cell.border = THIN_BORDER

    auto_width(ws)
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:N{len(rows) + 1}"
    return ws


def create_hr_detail_sheet(wb):
    """Sheet 3: HR Detail Log."""
    ws = wb.create_sheet("HR Details")

    headers = [
        "Date", "Player Name", "Player ID", "Bats", "Team",
        "Opponent", "Home/Away", "Pitcher Name", "Pitcher ID",
        "Pitcher Hand", "Batter Side", "Inning", "RBI",
        "Launch Speed", "Launch Angle", "Distance", "Description",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, HEADER_FILL_GREEN, len(headers))

    csv_headers, rows = load_csv("hr_details.csv")
    float_cols = {13, 14, 15}  # launch speed, angle, distance
    int_cols = {2, 8, 12}  # player_id, pitcher_id, rbi
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row):
            cell = ws.cell(row=i, column=j + 1)
            if j in float_cols:
                try:
                    cell.value = float(val) if val else ""
                except ValueError:
                    cell.value = val
            elif j in int_cols:
                try:
                    cell.value = int(val) if val else ""
                except ValueError:
                    cell.value = val
            else:
                cell.value = val
            cell.border = THIN_BORDER

    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Q{len(rows) + 1}"
    return ws


def create_odds_sheet(wb):
    """Sheet 4: Odds Tracking (template with headers, no data yet)."""
    ws = wb.create_sheet("Odds Tracking")

    headers = [
        "Date", "Player Name", "Opponent", "Home/Away",
        "Pitcher (Probable)", "Pitcher Hand",
        "DraftKings HR Odds", "FanDuel HR Odds", "BetMGM HR Odds",
        "Best Odds", "Implied Prob", "Did Hit HR?",
        "Season HR Rate", "Situation HR Rate", "Edge",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, HEADER_FILL_ORANGE, len(headers))

    # Add example formulas in row 2
    ws.cell(row=2, column=10, value='=MAX(G2,H2,I2)')  # Best Odds (for + odds)
    ws.cell(row=2, column=11, value='=IF(J2>0,100/(J2+100),ABS(J2)/(ABS(J2)+100))')  # Implied prob
    ws.cell(row=2, column=15, value='=IF(AND(N2<>"",K2<>""),N2-K2,"")')  # Edge

    # Add notes row
    notes = [
        "YYYY-MM-DD", "From Roster", "Team abbr", "Home/Away",
        "Check lineup", "L or R",
        "American odds", "American odds", "American odds",
        "Formula", "Formula", "Y/N (fill after game)",
        "From Roster", "Calc from splits", "Formula",
    ]
    for col, note in enumerate(notes, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = note
        cell.font = Font(italic=True, color="808080", size=9)

    auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


def create_analysis_sheet(wb):
    """Sheet 5: Analysis Dashboard with summary formulas and pivot-like layout."""
    ws = wb.create_sheet("Analysis")

    # Section 1: Player Summary
    ws.cell(row=1, column=1, value="PLAYER SITUATIONAL HR RATES")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E79")

    summary_headers = [
        "Player", "Bats", "Total HRs", "Games",
        "HR Rate", "Home HRs", "Home Games", "Home HR Rate",
        "Away HRs", "Away Games", "Away HR Rate",
        "vs LHP HRs", "vs LHP ABs", "vs LHP HR/AB",
        "vs RHP HRs", "vs RHP ABs", "vs RHP HR/AB",
    ]
    for col, h in enumerate(summary_headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, HEADER_FILL_PURPLE, len(summary_headers))

    # Load splits data
    csv_headers, rows = load_csv("splits.csv")
    for i, row in enumerate(rows, 4):
        # row format: name, id, team, league, bats, home_games, home_ab, home_hr, home_avg, home_ops,
        #             away_games, away_ab, away_hr, away_avg, away_ops,
        #             vs_lhp_games, vs_lhp_ab, vs_lhp_hr, vs_lhp_avg, vs_lhp_ops,
        #             vs_rhp_games, vs_rhp_ab, vs_rhp_hr, vs_rhp_avg, vs_rhp_ops
        name = row[0]
        bats = row[4]
        home_games = int(row[5]) if row[5] else 0
        home_ab = int(row[6]) if row[6] else 0
        home_hr = int(row[7]) if row[7] else 0
        away_games = int(row[10]) if row[10] else 0
        away_ab = int(row[11]) if row[11] else 0
        away_hr = int(row[12]) if row[12] else 0
        vlhp_games = int(row[15]) if row[15] else 0
        vlhp_ab = int(row[16]) if row[16] else 0
        vlhp_hr = int(row[17]) if row[17] else 0
        vrhp_games = int(row[20]) if row[20] else 0
        vrhp_ab = int(row[21]) if row[21] else 0
        vrhp_hr = int(row[22]) if row[22] else 0

        total_hr = home_hr + away_hr
        total_games = home_games + away_games

        values = [
            name, bats, total_hr, total_games,
            total_hr / total_games if total_games > 0 else 0,
            home_hr, home_games, home_hr / home_games if home_games > 0 else 0,
            away_hr, away_games, away_hr / away_games if away_games > 0 else 0,
            vlhp_hr, vlhp_ab, vlhp_hr / vlhp_ab if vlhp_ab > 0 else 0,
            vrhp_hr, vrhp_ab, vrhp_hr / vrhp_ab if vrhp_ab > 0 else 0,
        ]
        for j, val in enumerate(values):
            cell = ws.cell(row=i, column=j + 1)
            cell.value = val
            cell.border = THIN_BORDER
            if j in (4, 7, 10, 13, 16):  # Rate columns
                cell.number_format = '0.000'

    # Section 2: Key Insights area
    insight_row = len(rows) + 6
    ws.cell(row=insight_row, column=1, value="KEY OBSERVATIONS")
    ws.cell(row=insight_row, column=1).font = Font(bold=True, size=14, color="1F4E79")

    notes = [
        "Note: Early season sample sizes are small. Look for patterns as data accumulates.",
        "Target: 50+ games per player before drawing wagering conclusions.",
        "Edge = (Actual HR Rate in situation) - (Implied Probability from odds).",
        "Positive edge = potential value bet. Track results to validate.",
        "",
        "SITUATIONAL PATTERNS TO WATCH:",
        "- Lefty hitters vs RHP at home (typically highest HR rate scenario)",
        "- Righty hitters vs LHP (platoon advantage)",
        "- Switch hitters — compare their L vs R side performance",
        "- Home park factors (COL, CIN, PHI tend to boost HRs)",
        "- Pitcher hand matchup is often more predictive than home/away",
    ]
    for k, note in enumerate(notes):
        ws.cell(row=insight_row + 1 + k, column=1, value=note)
        ws.cell(row=insight_row + 1 + k, column=1).font = Font(
            italic=True if k == 0 else False, size=10
        )

    auto_width(ws, min_width=12)
    ws.freeze_panes = "A4"
    return ws


def main():
    print("Creating HR Tracker workbook...")
    wb = openpyxl.Workbook()

    create_roster_sheet(wb)
    print("  Created: Roster")

    create_daily_log_sheet(wb)
    print("  Created: Daily Game Log")

    create_hr_detail_sheet(wb)
    print("  Created: HR Details")

    create_odds_sheet(wb)
    print("  Created: Odds Tracking")

    create_analysis_sheet(wb)
    print("  Created: Analysis")

    wb.save(OUTPUT_FILE)
    print(f"\nWorkbook saved: {OUTPUT_FILE}")
    print("Upload this file to Google Drive to use as a Google Sheet.")


if __name__ == "__main__":
    main()
